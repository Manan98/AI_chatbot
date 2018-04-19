from openpyxl import load_workbook

wb = load_workbook('qanda.xlsx')
ws = wb['Sheet2']

sentences = []
j=-1
for i in range(2, 3998):
    try:
        j+=1
        sentences.append('L' + (str)(j) + ' +++$+++ u0 +++$+++ m0 +++$+++ BIANCA +++$+++ '+ws.cell(row=i,column=1).value)
    except:
        sentences.append('L' + (str)(j) + ' +++$+++ u0 +++$+++ m0 +++$+++ BIANCA +++$+++ ')
        pass
    try:
        j+=1
        sentences.append('L' + (str)(j) + ' +++$+++ u1 +++$+++ m0 +++$+++ CAMERON +++$+++ ' + ws.cell(row=i, column=2).value)
    except:
        sentences.append('L' + (str)(j) + ' +++$+++ u1 +++$+++ m0 +++$+++ CAMERON +++$+++')
        pass

table1 = open('table1.txt', 'w')
for item in sentences:
  table1.write("%s\n" % item)
